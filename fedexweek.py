from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook, Workbook
import re
import time
import os
from datetime import datetime

# =================  =================
# FedEx 網址
url = "https://www.fedex.com/en-us/shipping/fuel-surcharge.html"

# 主檔案 (要改公式的檔案)
calculator_path = r"C:\Users\BSM04\Desktop\FeeCalculator-fedex_20250730.xlsx"
sheet_name = "FedEx Fee Calculator"

# 紀錄檔案 (新創的 Excel，用來紀錄歷史，放在桌面)
log_path = r"C:\Users\BSM04\Desktop\FedEx_Update_Log.xlsx"
# =========================================================

def get_fedex_rate():
    """抓取 FedEx 網頁上的費率"""
    options = Options()
    options.add_argument("--disable-blink-features=AutomationControlled") 
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    options.add_argument("--start-maximized") 
    
    # 阻擋瀏覽器跳出來
    # options.add_argument("--headless") 

    print("[1/3] 正在啟動瀏覽器抓取費率...")
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

    try:
        driver.get(url)
        time.sleep(8) # 等待載入
        content = driver.page_source
        
        # 抓取數字 
        pattern = r"(\d{1,2}\.\d{2})%"
        matches = re.findall(pattern, content)

        if matches:
            target_rate = matches[0] 
            print(f"成功抓取最新費率: {target_rate}%")
            return target_rate
        else:
            print("網頁中找不到費率數字")
            return None
    except Exception as e:
        print(f"抓取錯誤: {e}")
        return None
    finally:
        driver.quit() 

def update_calculator(rate):
    """任務一：更新計算表的公式"""
    print(f"📂[2/3] 正在更新計算表: {os.path.basename(calculator_path)}")
    
    if not os.path.exists(calculator_path):
        print(f"❌找不到計算表檔案：{calculator_path}")
        return False

    try:
        wb = load_workbook(calculator_path)
        if sheet_name not in wb.sheetnames:
            print(f"❌找不到分頁 '{sheet_name}'")
            return False
            
        sheet = wb[sheet_name]
        new_formula = f'=VLOOKUP({rate}%,$Q$1:$T$37,4,0)'
        
        k_column_index = 11 
        count = 0
        
        # 更新每一列
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value is not None:
                sheet.cell(row=row, column=k_column_index).value = new_formula
                count += 1
        
        wb.save(calculator_path)
        print(f"   └── 已更新 {count} 筆公式為: {new_formula}")
        return True

    except PermissionError:
        print("❌計算表存檔失敗！檔案正被開啟中，請關閉後再試。")
        return False
    except Exception as e:
        print(f"計算表更新錯誤: {e}")
        return False

def update_log(rate):
    """任務二：寫入歷史紀錄檔"""
    print(f"📝 [3/3] 正在寫入紀錄檔: {os.path.basename(log_path)}")

    try:
        # 檢查紀錄檔是否存在
        if os.path.exists(log_path):
            wb = load_workbook(log_path)
            sheet = wb.active
        else:
            print("   └── 紀錄檔不存在，正在建立新檔案...")
            wb = Workbook()
            sheet = wb.active
            sheet.title = "Run History"
            # 建立標題列
            sheet.append(["執行時間", "抓取到的費率 (%)", "狀態"])

        # 準備資料
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 寫入一行新資料
        sheet.append([current_time, float(rate), "成功更新"])
        
        wb.save(log_path)
        print(f"   └── ✅ 紀錄已儲存！(時間: {current_time}, 費率: {rate}%)")

    except PermissionError:
        print("❌ 紀錄檔存檔失敗！檔案正被開啟中，請關閉。")
    except Exception as e:
        print(f"❌ 寫入紀錄錯誤: {e}")

# ================= 主程式執行區 =================
if __name__ == "__main__":
    print("=== 程式開始執行 ===")
    
    # 1. 抓取費率
    rate = get_fedex_rate()
    
    if rate:
        # 2. 更新主計算表
        success = update_calculator(rate)
        
        # 3. 只有在抓取成功時，才寫入紀錄
        if success:
            update_log(rate)
        else:
            print("⚠️ 計算表更新失敗，跳過寫入紀錄。")
    
    print("=== 程式執行結束 ===")